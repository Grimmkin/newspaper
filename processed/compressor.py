import json
import os
import glob
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment
from datetime import datetime

# List of states
states = [
    "Abia", "Adamawa", "Akwa Ibom", "Anambra", "Bauchi", "Bayelsa", "Benue", "Borno", "Cross River", "Delta",
    "Ebonyi", "Edo", "Ekiti", "Enugu", "Federal Capital Territory", "FCT", "Gombe", "Imo", "Jigawa", "Kaduna",
    "Kano", "Katsina", "Kebbi", "Kogi", "Kwara", "Lagos", "Nasarawa", "Niger", "Ogun", "Ondo", "Osun", "Oyo",
    "Plateau", "Rivers", "Sokoto", "Taraba", "Yobe", "Zamfara", "NO LOCATION"
]

# Create a new workbook for regular states
wb_regular = Workbook()
ws_regular = wb_regular.active

# Set column headers for regular states
headers_regular = ["URL", "Summary", "State"]
for col_num, header in enumerate(headers_regular, 1):
    cell = ws_regular.cell(row=1, column=col_num)
    cell.value = header
    cell.alignment = Alignment(horizontal="center")

# Create a new workbook for "NO LOCATION"
wb_no_location = Workbook()
ws_no_location = wb_no_location.active

# Set column headers for "NO LOCATION"
headers_no_location = ["URL", "Summary"]
for col_num, header in enumerate(headers_no_location, 1):
    cell = ws_no_location.cell(row=1, column=col_num)
    cell.value = header
    cell.alignment = Alignment(horizontal="center")

# Set column widths for regular states
column_widths_regular = [30, 60, 15]
for col_num, column_width in enumerate(column_widths_regular, 1):
    column_letter = get_column_letter(col_num)
    ws_regular.column_dimensions[column_letter].width = column_width

# Set column widths for "NO LOCATION"
column_widths_no_location = [30, 60]
for col_num, column_width in enumerate(column_widths_no_location, 1):
    column_letter = get_column_letter(col_num)
    ws_no_location.column_dimensions[column_letter].width = column_width

# Read JSON files in the directory
json_files = glob.glob("*.json")

regular_data = []  # List to store regular state data
no_location_data = []  # List to store "NO LOCATION" data

for file_path in json_files:
    with open(file_path, "r") as file:
        file_data = json.load(file)

    for state in states:
        if state in file_data:
            articles = file_data[state]
            for article in articles:
                if state in ["FCT", "Federal Capital Territory", "Abuja"]:
                    state = "Federal Capital Territory"
                if state == "NO LOCATION":
                    no_location_data.append((article["url"], article["summary"]))
                else:
                    regular_data.append((article["url"], article["summary"], state))

# Sort the regular state data alphabetically by the state column
sorted_regular_data = sorted(regular_data, key=lambda x: x[2])

# Populate the regular state worksheet with the sorted data
for row_data in sorted_regular_data:
    ws_regular.append(row_data)

# Populate the "NO LOCATION" worksheet
for row_data in no_location_data:
    ws_no_location.append(row_data)

# Save the regular state workbook with the current date as the filename
output_filename_regular = f"data_compilation_regular_{datetime.now().strftime('%Y-%m-%d')}.xlsx"
wb_regular.save(output_filename_regular)

# Save the "NO LOCATION" workbook with the current date as the filename
output_filename_no_location = f"data_compilation_no_location_{datetime.now().strftime('%Y-%m-%d')}.xlsx"
wb_no_location.save(output_filename_no_location)

print(f"Regular state data compilation saved as {output_filename_regular}")
print(f"'NO LOCATION' data compilation saved as {output_filename_no_location}")