import json
import os
import glob
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment
from datetime import datetime

# List of states
states = [
    "Abia", "Adamawa", "Akwa Ibom", "Anambra", "Bauchi", "Bayelsa", "Benue", "Borno", "Cross River", "Delta",
    "Ebonyi", "Edo", "Ekiti", "Enugu", "Federal Capital Territory", "Gombe", "Imo", "Jigawa", "Kaduna",
    "Kano", "Katsina", "Kebbi", "Kogi", "Kwara", "Lagos", "Nasarawa", "Niger", "Ogun", "Ondo", "Osun", "Oyo",
    "Plateau", "Rivers", "Sokoto", "Taraba", "Yobe", "Zamfara", "NO LOCATION"
]

# Create a new workbook for regular states
wb_regular = Workbook()
ws_regular = wb_regular.active

# Set column headers for "NO LOCATION"
headers_regular = ["State", "URL"]
for col_num, header in enumerate(headers_regular, 1):
    cell = ws_regular.cell(row=1, column=col_num)
    cell.value = header
    cell.alignment = Alignment(horizontal="center")

# Set column widths for "NO LOCATION"
column_widths_regular = [30, 60]
for col_num, column_width in enumerate(column_widths_regular, 1):
    column_letter = get_column_letter(col_num)
    ws_regular.column_dimensions[column_letter].width = column_width

regular_data = []  # List to store regular state data

with open('articles_by_state.json', "r") as file:
    file_data = json.load(file)

for state in states:
    for i in file_data[state]:
        if state in ["FCT", "Federal Capital Territory", "Abuja"]:
            state = "Federal Capital Territory"
        regular_data.append((state, i))

# Sort the regular state data alphabetically by the state column
sorted_regular_data = sorted(regular_data, key=lambda x: x[0])

# Populate the regular state worksheet with the sorted data
for row_data in sorted_regular_data:
    ws_regular.append(row_data)

# Save the regular state workbook with the current date as the filename
output_filename_regular = f"data_compilation_regular_{datetime.now().strftime('%Y-%m-%d')}.xlsx"
wb_regular.save(output_filename_regular)

print(f"Regular state data compilation saved as {output_filename_regular}")